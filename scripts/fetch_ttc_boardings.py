"""Build a small committed REAL slice of TTC subway boardings (demo_data/).

Produces ``demo_data/ttc_boardings__downtown.csv`` — real typical-weekday boardings for
the downtown-core subway stations, the magnitude behind an honest "measured TTC boardings"
source for the Urban-OS TransitLoad lens (ADR-0031).

Source (real, only reshaped): the City "TTC Ridership - Subway-Scarborough RT Station
Usage" dataset (``ttc-ridership-subway-scarborough-rt-station-usage``), latest year. Its
XLSX lists, per station, the typical-day "Total" customers travelling to/from the platforms
(summed here across the lines a station serves). We keep the downtown-core stations that the
demo substrate models and tag each with its **real station coordinate**.

Why a built-in coord table (not GTFS): the TTC GTFS feed exposes no clean subway-*station*
coordinates (no parent stations; only mixed platform/surface stops needing a route_type
join), so we place each station at its known real location — the same verified coordinates
the substrate uses. The *boardings are real measured values*; only the intraday distribution
is modelled later (in ``adapters.ttc_boardings_by_node``), and that boundary is stated there.

Honesty: the committed slice is **pure real data** — a daily total per station, no modelled
shape baked in. Offline-safe: openpyxl missing or any network/parse failure prints a note and
exits 0 (the adapter's synthetic fallback covers dev/CI), so ``make demo-data`` never breaks.
The raw XLSX is never committed — only the small normalized slice.

    python scripts/fetch_ttc_boardings.py
"""
from __future__ import annotations

import csv
import sys
import tempfile
from pathlib import Path

# Allow running as a plain script without installing the package.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from urbanos.risk.ingest.ckan import CKANClient  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "demo_data"
SLUG = "ttc-ridership-subway-scarborough-rt-station-usage"

# Downtown-core subway stations the demo substrate models -> their REAL coordinates
# (the verified station locations also used by adapters.toronto._NODES). Station-usage
# names are matched to these after normalization (uppercase, strip punctuation).
STATION_COORDS: dict[str, tuple[float, float]] = {
    "UNION": (43.6452, -79.3806),
    "KING": (43.6489, -79.3777),
    "QUEEN": (43.6526, -79.3793),
    "ST ANDREW": (43.6479, -79.3849),
    "ST PATRICK": (43.6549, -79.3884),
    "OSGOODE": (43.6505, -79.3866),
}


def _norm(name: str) -> str:
    """Normalize a station name for matching: uppercase, drop punctuation, collapse spaces."""
    out = []
    for ch in str(name).upper():
        out.append(ch if ch.isalnum() or ch == " " else " ")
    return " ".join("".join(out).split())


def _station_totals(xlsx_path: Path) -> dict[str, float]:
    """Parse the station-usage XLSX into {NORMALIZED_STATION: total_boardings}, summing the
    'Total' column across the rows for a station (interchange stations appear once per line)."""
    import openpyxl  # fetch-only dependency; guarded by the caller

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()  # read_only keeps the file open on Windows; release before temp cleanup
    # Find the header row (has 'Station' and 'Total') and its column indices.
    station_col = total_col = None
    data_start = 0
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "station" in cells and "total" in cells:
            station_col = cells.index("station")
            total_col = cells.index("total")
            data_start = i + 1
            break
    if station_col is None or total_col is None:
        raise RuntimeError("station-usage sheet: no 'Station'/'Total' header found")
    totals: dict[str, float] = {}
    for row in rows[data_start:]:
        if station_col >= len(row) or total_col >= len(row):
            continue
        name, total = row[station_col], row[total_col]
        if not name or not isinstance(total, (int, float)):
            continue
        totals[_norm(name)] = totals.get(_norm(name), 0.0) + float(total)
    return totals


def main() -> int:
    def _year(name: str) -> int:
        years = [int(t) for t in "".join(c if c.isdigit() else " " for c in name).split()
                 if len(t) == 4 and t.startswith("20")]
        return max(years) if years else 0

    try:
        with CKANClient() as ckan:
            xlsx = [r for r in ckan.resources(SLUG)
                    if (r.get("format") or "").lower() == "xlsx"]
            res = max(xlsx, key=lambda r: _year(r.get("name", "")), default=None)
            if res is None:
                print("fetch_ttc_boardings: no XLSX resource found; leaving slice in place.")
                return 0
            with tempfile.TemporaryDirectory() as td:
                xlsx = ckan.download_resource(res, Path(td) / "usage.xlsx", max_bytes=20_000_000)
                totals = _station_totals(xlsx)
    except ImportError:
        print("fetch_ttc_boardings: openpyxl not installed (fetch-only dep: pip install "
              "openpyxl). The synthetic fallback covers dev/CI.")
        return 0
    except Exception as exc:  # noqa: BLE001 — offline-safe: never fail the chained target
        print(f"fetch_ttc_boardings: skipped (no network / API / parse error: {exc}). "
              "The synthetic fallback covers dev/CI.")
        return 0

    rows_out: list[dict] = []
    matched: list[str] = []
    for station, (lat, lng) in STATION_COORDS.items():
        boardings = totals.get(station)
        if boardings is None:
            continue
        matched.append(station)
        rows_out.append({
            "location": station, "lat": lat, "lng": lng,
            "mode": "subway", "boardings": int(round(boardings)),
        })
    if not rows_out:
        print("fetch_ttc_boardings: no downtown stations matched; leaving slice in place.")
        return 0

    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / "ttc_boardings__downtown.csv"
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["location", "lat", "lng", "mode", "boardings"])
        w.writeheader()
        w.writerows(rows_out)
    print(f"ttc_boardings: {len(rows_out)} downtown stations {matched} -> {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
